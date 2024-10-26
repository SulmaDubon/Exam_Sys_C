# admin_panel/views.py
import random
from django.forms import ValidationError, formset_factory
from django.shortcuts import render, get_object_or_404, redirect
from django.views import View
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.urls import reverse_lazy
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils.decorators import method_decorator
import openpyxl
import pandas as pd
from users.models import CustomUser
from dashboard_users.models import Examen, InscripcionExamen, Modulo, Pregunta, Respuesta, TipoExamen, UserExam
from dashboard_users.forms import ExamenForm, PreguntaFormSet, RespuestaFormSet,  TipoExamenForm, ModuloFormSet, SubirPreguntasForm, PreguntaForm  # Importar ExamenForm desde admin_panel/forms.py
from users.forms import UserRegistrationForm  # Importar UserRegistrationForm desde users/forms.py
from django.contrib.auth.views import LoginView
from django.contrib import messages 
from datetime import date, datetime, timedelta
from django.core.paginator import Paginator
from django.db.models import Q
from django.urls import reverse
from django.db.models import Max
from django.views.generic.edit import FormView
from django.http import HttpResponse, JsonResponse
from django.db import transaction 
from django.db.models import Prefetch 
from openpyxl.styles import PatternFill 
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.units import inch


# from .decorators import es_admin


#-------------------------------------------------------
# Función para verificar si el usuario es administrador
#-------------------------------------------------------
def es_admin(usuario):
    return usuario.is_superuser

#-----------------------------------------------------------
# Login Administrador
#------------------------------------------------------------

#@method_decorator(user_passes_test(es_admin), name='dispatch')
class AdminLoginView(LoginView):
    template_name = 'admin_panel/admin_login.html'

    def get_success_url(self):
        return reverse_lazy('admin_panel:admin_panel')

#-----------------------------------------------------------------
#    ADMIN PANEL
#----------------------------------------------------------------

# Vista principal del panel de administración
@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class VistaAdminPanel(View):
    def get(self, request):
        # Renderiza la plantilla principal del panel de administración
        return render(request, 'admin_panel/admin_panel.html')


#-------------------------------------------------------------------
#                 USUARIOS
#--------------------------------------------------------------------

# Vista para listar todos los usuarios
@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class ListaUsuarios(ListView):
    model = CustomUser
    template_name = 'admin_panel/lista_usuarios.html'
    context_object_name = 'usuarios'
    # Renderiza la plantilla con la lista de usuarios

# Vista para crear un nuevo usuario
@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class CrearUsuario(CreateView):
    model = CustomUser
    form_class = UserRegistrationForm
    template_name = 'admin_panel/formulario_usuario.html'
    success_url = reverse_lazy('admin_panel:lista_usuarios')
    # Renderiza el formulario para crear un usuario y maneja su creación

# Vista para editar un nuevo usuario
@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class EditarUsuario(UpdateView):
    model = CustomUser
    form_class = UserRegistrationForm
    template_name = 'admin_panel/formulario_usuario.html'
    success_url = reverse_lazy('admin_panel:lista_usuarios')

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Usuario actualizado exitosamente.')
        return response

    def form_invalid(self, form):
        messages.error(self.request, 'Hubo un error al actualizar el usuario. Por favor, revisa los datos ingresados.')
        return super().form_invalid(form)

# Vista para eliminar un usuario existente
@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class EliminarUsuario(DeleteView):
    model = CustomUser
    template_name = 'admin_panel/confirmar_eliminacion_usuario.html'
    success_url = reverse_lazy('admin_panel:lista_usuarios')
    # Renderiza una página de confirmación y maneja la eliminación del usuario

#------------------------------------------------------------
#              LISTAR   EXAMEN
#-----------------------------------------------------------

# Vista para listar todos los exámenes
@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class ListaExamenes(ListView):
    model = Examen
    template_name = 'admin_panel/lista_examenes.html'
    context_object_name = 'examenes'
    paginate_by = 10  # Número de elementos por página

    def get_queryset(self):
        queryset = Examen.objects.all()
        order = self.request.GET.get('order', 'fecha')
        direction = self.request.GET.get('direction', 'asc')
        year = self.request.GET.get('year')
        month = self.request.GET.get('month')
        
        if year and month:
            queryset = queryset.filter(fecha__year=year, fecha__month=month)
        elif year:
            queryset = queryset.filter(fecha__year=year)
        elif month:
            queryset = queryset.filter(fecha__month=month)
        
        if direction == 'desc':
            order = '-' + order
        
        return queryset.order_by(order)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_year = datetime.now().year
        current_month = datetime.now().month

        # Obtener valores de los parámetros GET o asignar por defecto el año y mes actuales
        selected_year = self.request.GET.get('year', current_year)
        selected_month = self.request.GET.get('month', current_month)

        # Crear listas de años y meses para el filtrado
        year_list = list(range(current_year - 5, current_year + 5))
        month_list = [
            {'value': i, 'name': datetime(1900, i, 1).strftime('%B')}
            for i in range(1, 13)
        ]

        # Actualizar el contexto con los datos necesarios
        context['year_list'] = year_list
        context['month_list'] = month_list
        context['current_year'] = int(selected_year)  # Asegurarse de que sean enteros
        context['current_month'] = int(selected_month)

        return context

# Vista para crear un nuevo examen
@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class CrearExamen(CreateView):
    model = Examen
    form_class = ExamenForm
    template_name = 'admin_panel/formulario_examen.html'
    success_url = reverse_lazy('admin_panel:lista_examenes')

    def form_valid(self, form):
        response = super().form_valid(form)
        examen = self.object

        try:
            with transaction.atomic():  # Inicia una transacción
                preguntas_ids_seleccionadas = set()

                for modulo in examen.tipo_examen.modulos.all():
                    todas_las_preguntas = list(Pregunta.objects.filter(modulo=modulo).values_list('id', 'identificador_de_grupo', 'orden'))
                    random.shuffle(todas_las_preguntas)

                    total_preguntas_requeridas = modulo.cantidad_preguntas
                    total_preguntas_disponibles = len(todas_las_preguntas)

                    # Si no hay suficientes preguntas, ajusta al total disponible
                    if total_preguntas_disponibles < total_preguntas_requeridas:
                        total_preguntas_requeridas = total_preguntas_disponibles

                    preguntas_seleccionadas_modulo = 0
                    for pregunta_id, identificador_grupo, _ in todas_las_preguntas:
                        if preguntas_seleccionadas_modulo >= total_preguntas_requeridas:
                            break

                        # Añadir las preguntas del grupo completo, si aplica
                        if identificador_grupo:
                            preguntas_del_grupo = list(Pregunta.objects.filter(identificador_de_grupo=identificador_grupo).order_by('orden').values_list('id', flat=True))
                            if len(preguntas_ids_seleccionadas) + len(preguntas_del_grupo) <= total_preguntas_requeridas:
                                preguntas_ids_seleccionadas.update(preguntas_del_grupo)
                                preguntas_seleccionadas_modulo += len(preguntas_del_grupo)
                        else:
                            # Añadir la pregunta individual
                            preguntas_ids_seleccionadas.add(pregunta_id)
                            preguntas_seleccionadas_modulo += 1

                # Guardar los IDs de las preguntas seleccionadas en el examen
                if preguntas_ids_seleccionadas:
                    examen.preguntas.set(preguntas_ids_seleccionadas)
                else:
                    messages.error(self.request, 'Insuficientes preguntas para este examen. El examen se ha guardado, pero no tiene preguntas asignadas.')

                messages.success(self.request, f'Se han asignado {len(preguntas_ids_seleccionadas)} preguntas al examen.')

        except Exception as e:
            messages.error(self.request, f'Hubo un error al asignar preguntas: {str(e)}')
            return self.form_invalid(form)

        return response

    def form_invalid(self, form):
        messages.error(self.request, 'Hubo un error al crear el examen. Por favor, revisa los datos ingresados.')
        return super().form_invalid(form)

# Vista para editar un examen existente
@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class EditarExamen(UpdateView):
    model = Examen
    form_class = ExamenForm
    template_name = 'admin_panel/formulario_examen.html'
    success_url = reverse_lazy('admin_panel:lista_examenes')

# Vista para eliminar un examen existente
@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class EliminarExamen(DeleteView):
    model = Examen
    template_name = 'admin_panel/confirmar_eliminacion_examen.html'
    success_url = reverse_lazy('admin_panel:lista_examenes')


#------------------------------------------------------------------
#     USUARIOS POR EXAMEN
#-----------------------------------------------------------------

@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class UsuariosInscritosView(DetailView):
    model = Examen
    template_name = 'admin_panel/usuarios_inscritos.html'
    context_object_name = 'examen'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        search_query = self.request.GET.get('search', '')
        
        if search_query:
            usuarios = self.object.usuarios.filter(
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(cedula__icontains=search_query)
            )
        else:
            usuarios = self.object.usuarios.all()
        
        paginator = Paginator(usuarios, 20)  # Mostrar 20 usuarios por página
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context['search_query'] = search_query
        context['page_obj'] = page_obj
        return context

#-------------------------------------------------------
#           VISTA PARA MODULOS ASOCIADOS A EXAMEN
#--------------------------------------------------------

def get_modulos(request, tipo_examen_id):
    # Filtra los módulos por el tipo de examen seleccionado
    modulos = Modulo.objects.filter(tipo_examen_id=tipo_examen_id)
    data = {
        'modulos': [{'id': modulo.id, 'nombre': modulo.nombre} for modulo in modulos]
    }
    return JsonResponse(data)


#------------------------------------------------------------------------------
#      PREGUNTAS
#------------------------------------------------------------------------------
# listar pregunta simple
@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class ListaPreguntas(ListView):
    model = Pregunta
    template_name = 'admin_panel/lista_preguntas.html'
    context_object_name = 'preguntas'
    paginate_by = 10  # Número de preguntas por página

    def get_queryset(self):
        # Obtener todas las preguntas con sus respuestas relacionadas
        return Pregunta.objects.prefetch_related('respuestas').order_by('id')


@method_decorator([login_required, user_passes_test(es_admin)], name='dispatch')
class CrearPreguntaView(View):
    template_name = 'admin_panel/formulario_preguntas.html'
    success_url = reverse_lazy('admin_panel:lista_preguntas')

    def get(self, request, *args, **kwargs):
        form = PreguntaForm()  # Formulario para atributos comunes
        formset = PreguntaFormSet(queryset=Pregunta.objects.none())  # Formset para múltiples preguntas
        return render(request, self.template_name, {'form': form, 'formset': formset})

    def post(self, request, *args, **kwargs):
        form = PreguntaForm(request.POST)
        formset = PreguntaFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            es_agrupada = request.POST.get('es_agrupada')  # Tomamos si el usuario quiere agrupar
            modulo = request.POST.get('modulo')
            tipo_examen = request.POST.get('tipo_examen')

            if es_agrupada:
                # Crear un nuevo grupo
                ultimo_grupo = Pregunta.objects.filter(identificador_de_grupo__isnull=False).order_by('-identificador_de_grupo').first()
                if ultimo_grupo:
                    identificador_grupo = ultimo_grupo.identificador_de_grupo + 1
                else:
                    identificador_grupo = 1  # Si no existen grupos, empezar con 1

                # Crear cada pregunta dentro del nuevo grupo
                for index, pregunta_form in enumerate(formset):
                    if pregunta_form.cleaned_data.get('texto'):
                        Pregunta.objects.create(
                            texto=pregunta_form.cleaned_data['texto'],
                            modulo=modulo,
                            tipo_examen=tipo_examen,
                            identificador_de_grupo=identificador_grupo,
                            orden=index + 1,
                            activo=True
                        )
            else:
                # Crear cada pregunta como individual
                for pregunta_form in formset:
                    if pregunta_form.cleaned_data.get('texto'):
                        Pregunta.objects.create(
                            texto=pregunta_form.cleaned_data['texto'],
                            modulo=modulo,
                            tipo_examen=tipo_examen,
                            identificador_de_grupo=None,
                            orden=1,
                            activo=True
                        )

            messages.success(request, "Las preguntas han sido creadas con éxito.")
            return redirect(self.success_url)

        messages.error(request, "Hubo un error al crear las preguntas. Por favor, revisa los campos e inténtalo de nuevo.")
        return render(request, self.template_name, {'form': form, 'formset': formset})

# Eliminar pregunta
class PreguntaDeleteView(DeleteView):
    model = Pregunta
    template_name = 'admin_panel/confirmar_eliminacion_pregunta.html'
    success_url = reverse_lazy('admin_panel:lista_preguntas')

# subir pregunta
class SubirPreguntasView(FormView):
    template_name = 'admin_panel/subir_preguntas.html'
    form_class = SubirPreguntasForm
    success_url = reverse_lazy('admin_panel:lista_preguntas')

    def form_valid(self, form):
        archivo_excel = form.cleaned_data['archivo']
        modulo = form.cleaned_data['modulo']
        tipo_examen = form.cleaned_data['tipo_examen']

        try:
            wb = openpyxl.load_workbook(archivo_excel)
            sheet = wb.active

            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # Verificar si la fila está completamente vacía
                if all(cell is None for cell in row):
                    continue  # Ignorar filas completamente vacías

                # Obtener los valores de las columnas
                pregunta_texto = row[0]
                respuesta_correcta = row[1]
                respuesta_1 = row[2]
                respuesta_2 = row[3]
                respuesta_3 = row[4]
                identificador_grupo = row[5]

                # Validar que la pregunta no esté vacía (eliminando espacios en blanco primero)
                if not pregunta_texto or not str(pregunta_texto).strip():
                    mensajes_error = f"La fila {row_number} del archivo no tiene un texto de pregunta válido. Por favor, corrija el archivo e inténtelo de nuevo."
                    messages.error(self.request, mensajes_error)
                    return self.form_invalid(form)

                pregunta_texto = str(pregunta_texto).strip()

                # Determinar si la pregunta es parte de un grupo o es individual
                identificador_grupo = int(identificador_grupo) if identificador_grupo else None
                orden = 1
                if identificador_grupo:
                    # Si hay un identificador de grupo, calcular el siguiente orden disponible
                    preguntas_existentes = Pregunta.objects.filter(identificador_de_grupo=identificador_grupo)
                    orden = preguntas_existentes.count() + 1

                # Crear la pregunta
                pregunta_actual = Pregunta.objects.create(
                    texto=pregunta_texto,
                    modulo=modulo,
                    tipo_examen=tipo_examen,
                    identificador_de_grupo=identificador_grupo,
                    orden=orden,
                    activo=True
                )

                # Crear las respuestas asociadas a la pregunta
                respuestas = [
                    {'texto': str(respuesta_1).strip() if respuesta_1 else None, 'es_correcta': respuesta_1 == respuesta_correcta},
                    {'texto': str(respuesta_2).strip() if respuesta_2 else None, 'es_correcta': respuesta_2 == respuesta_correcta},
                    {'texto': str(respuesta_3).strip() if respuesta_3 else None, 'es_correcta': respuesta_3 == respuesta_correcta},
                ]

                for respuesta in respuestas:
                    if respuesta['texto']:  # Solo crear respuestas no vacías
                        Respuesta.objects.create(
                            pregunta=pregunta_actual,
                            texto=respuesta['texto'],
                            es_correcta=respuesta['es_correcta']
                        )

            messages.success(self.request, "Las preguntas se han subido correctamente.")
            return super().form_valid(form)

        except Exception as e:
            messages.error(self.request, f"Hubo un error al procesar el archivo: {str(e)}")
            return self.form_invalid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Hubo un error al subir las preguntas. Por favor, revisa el archivo e inténtalo nuevamente.")
        return self.render_to_response(self.get_context_data(form=form))


# Actualizar preguntas
class ActualizarPreguntasView(View):
    def get(self, request, *args, **kwargs):
        # Ejecutar la función de sincronización manual
        hoy = date.today()
        for examen in Examen.objects.all():
            if examen.fecha < hoy:
                continue

            # Verificar si realmente hay preguntas para sincronizar
            for user_exam in UserExam.objects.filter(examen=examen, finalizado=False):
                if user_exam.preguntas.exists():
                    user_exam.preguntas.set(examen.preguntas.all())
                    user_exam.save()

        # Mensaje de éxito para el usuario solo una vez
        messages.success(request, 'Las preguntas han sido actualizadas correctamente.')

        # Redirigir a la lista de preguntas
        return redirect(reverse('admin_panel:lista_preguntas'))

#------------------------------------------------
#   ACCIONES
#------------------------------------------------
class AccionesExamenesView(View):
    def post(self, request, *args, **kwargs):
        action = request.POST.get('action')
        selected_exams = request.POST.getlist('selected_exams')

        if not selected_exams:
            messages.warning(request, "No se seleccionaron exámenes.")
            return redirect('admin_panel:lista_examenes')

        if action == 'edit':
            if len(selected_exams) > 1:
                messages.warning(request, "Solo puedes editar un examen a la vez.")
                return redirect('admin_panel:lista_examenes')
            return redirect(reverse('admin_panel:editar_examen', args=[selected_exams[0]]))
        
        elif action == 'delete':
            for exam_id in selected_exams:
                exam = get_object_or_404(Examen, id=exam_id)
                exam.delete()
            messages.success(request, "Exámenes eliminados con éxito.")
        
        elif action == 'users':
            if len(selected_exams) > 1:
                messages.warning(request, "Solo puedes ver los usuarios inscritos de un examen a la vez.")
                return redirect('admin_panel:lista_examenes')
            return redirect(reverse('admin_panel:usuarios_inscritos', args=[selected_exams[0]]))

        return redirect('admin_panel:lista_examenes')


#---------------------------------------------
#             CREAR TIPO EXAMEN
#----------------------------------------------

class TipoExamenListView(ListView):
    model = TipoExamen
    template_name = 'admin_panel/lista_tipo_examen.html'
    context_object_name = 'tipos_examenes'

    def get_queryset(self):
        # Utilizamos prefetch_related para traer todos los módulos relacionados en una sola operación
        return TipoExamen.objects.prefetch_related('modulos')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Preparar información de módulos para cada tipo de examen
        tipos_examenes_info = []
        for tipo_examen in context['tipos_examenes']:
            # `tipo_examen.modulos.all()` usa los módulos prefetched gracias a `prefetch_related`
            modulos = tipo_examen.modulos.all()
            total_preguntas = sum(modulo.cantidad_preguntas for modulo in modulos)
            
            tipos_examenes_info.append({
                'tipo_examen': tipo_examen,
                'modulos': modulos,
                'total_preguntas': total_preguntas
            })

        # Añadir la información enriquecida al contexto
        context['tipos_examenes_info'] = tipos_examenes_info
        return context



class CrearTipoExamenView(View):
    def get(self, request):
        tipo_examen_form = TipoExamenForm()
        modulo_formset = ModuloFormSet()
        context = {
            'tipo_examen_form': tipo_examen_form,
            'modulo_formset': modulo_formset
        }
        return render(request, 'admin_panel/crear_tipo_examen.html', context)

    def post(self, request):
        tipo_examen_form = TipoExamenForm(request.POST)
        modulo_formset = ModuloFormSet(request.POST)

        if tipo_examen_form.is_valid() and modulo_formset.is_valid():
            tipo_examen = tipo_examen_form.save()
            
             # Vincular los módulos al tipo de examen
            modulo_formset.instance = tipo_examen
            modulo_formset.save()
            return redirect('admin_panel:lista_tipo_examen')  # Cambia 'ruta_donde_redirigir' a la URL de listar

        context = {
            'tipo_examen_form': tipo_examen_form,
            'modulo_formset': modulo_formset
        }
        return render(request, 'admin_panel/crear_tipo_examen.html', context)




class EditarTipoExamenView(View):
    def get(self, request, pk):
        tipo_examen = get_object_or_404(TipoExamen, pk=pk)
        tipo_examen_form = TipoExamenForm(instance=tipo_examen)
        modulo_formset = ModuloFormSet(instance=tipo_examen)
        context = {
            'tipo_examen_form': tipo_examen_form,
            'modulo_formset': modulo_formset
        }
        return render(request, 'editar_tipo_examen.html', context)

    def post(self, request, pk):
        tipo_examen = get_object_or_404(TipoExamen, pk=pk)
        tipo_examen_form = TipoExamenForm(request.POST, instance=tipo_examen)
        modulo_formset = ModuloFormSet(request.POST, instance=tipo_examen)

        if tipo_examen_form.is_valid() and modulo_formset.is_valid():
            tipo_examen = tipo_examen_form.save()
            modulo_formset.save()
            return redirect('ruta_donde_redirigir')

        context = {
            'tipo_examen_form': tipo_examen_form,
            'modulo_formset': modulo_formset
        }
        return render(request, 'editar_tipo_examen.html', context)
    

class TipoExamenDeleteView(DeleteView):
    model = TipoExamen
    template_name = 'admin_panel/confirmar_eliminar_tipo_examen.html'
    success_url = reverse_lazy('admin_panel:listar_tipo_examen')  # Redirige a la lista de exámenes después de eliminar


#-----------------------------------------------
#         RESULTADOS
#---------------------------------------------
class ResultadosAdministrador(View):
    def get(self, request, examen_id):
        examen = get_object_or_404(Examen, pk=examen_id)
        # Aquí puedes obtener los resultados asociados al examen si es necesario
        return render(request, 'admin_panel/resultados_administrador.html', {'examen': examen})
    
#-----------------------------------------------
#    INFORMES
#----------------------------------------------

def generar_informe_examen(request):
    # Obtener todos los exámenes
    examenes = Examen.objects.prefetch_related(
        Prefetch('preguntas', queryset=Pregunta.objects.all())
    )

    # Preparar una lista para almacenar datos del informe
    informe_data = []
    usuarios = []  # Lista para almacenar nombres de usuarios

    # Iterar sobre los exámenes
    for examen in examenes:
        inscripciones = InscripcionExamen.objects.filter(examen=examen)

        # Inicializar el diccionario para cada pregunta en el examen
        for pregunta in examen.preguntas.all():
            respuesta_correcta = pregunta.respuestas.filter(es_correcta=True).first()
            correcta_letra = respuesta_correcta.letra if respuesta_correcta else None

            # Crear la fila base para cada pregunta
            fila_base = {
                'Modulo': pregunta.modulo.nombre,
                'Pregunta': pregunta.id,
                'Respuesta Correcta': correcta_letra,
                'Aciertos': 0,
                'Errores': 0,
                'Nulas': 0,
            }

            # Procesar cada inscripción
            for inscripcion in inscripciones:
                try:
                    user_exam = UserExam.objects.get(usuario=inscripcion.usuario, examen=examen)
                    respuestas_usuario = user_exam.respuestas  # Obtener las respuestas del usuario

                    # Obtener la respuesta del usuario para la pregunta actual
                    respuesta_id = respuestas_usuario.get(str(pregunta.id), None)

                    # Obtener la letra correspondiente a la respuesta ID
                    letra_respuesta = None
                    if respuesta_id is not None:
                        respuesta = Respuesta.objects.get(id=respuesta_id)  # Obtén la respuesta usando el ID
                        letra_respuesta = respuesta.letra  # Almacena la letra

                    # Almacenar la respuesta del usuario en la fila
                    fila_base[inscripcion.usuario.username] = letra_respuesta  # Almacenar la letra de la respuesta

                    # Contar aciertos, errores y nulas
                    if letra_respuesta == correcta_letra:
                        fila_base['Aciertos'] += 1
                    elif letra_respuesta is None:  # Si no hay respuesta dada
                        fila_base['Nulas'] += 1
                    else:
                        fila_base['Errores'] += 1

                    # Agregar el usuario a la lista si no está ya
                    if inscripcion.usuario.username not in usuarios:
                        usuarios.append(inscripcion.usuario.username)

                except UserExam.DoesNotExist:
                    continue  # Si no existe, simplemente continúa con el siguiente
                except Respuesta.DoesNotExist:
                    # Manejo en caso de que el ID de respuesta no exista
                    fila_base[inscripcion.usuario.username] = None  # Marca como nulo si no se encuentra

            # Añadir la fila base a los datos del informe
            informe_data.append(fila_base)

    # Crear un DataFrame para los datos del informe
    df_informe = pd.DataFrame(informe_data)

    # Crear un nuevo archivo Excel para guardar el informe
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Informe de Respuestas"

    # Escribir encabezados
    headers = ['Modulo', 'Pregunta', 'Respuesta Correcta'] + usuarios + ['Aciertos', 'Errores', 'Nulas']
    ws.append(headers)

    # Escribir los datos
    for index, row in df_informe.iterrows():
        # Convertir respuestas None a espacios en blanco para el Excel
        row = [cell if cell is not None else '' for cell in row]
        ws.append(row)

    # Guardar el archivo
    output_file_path = 'informe_respuestas.xlsx'  # Ajusta la ruta según tu estructura
    wb.save(output_file_path)

    # Devolver el archivo como descarga
    with open(output_file_path, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{output_file_path}"'
        return response


def certificado(request, user_exam_id):
    # Obtener el UserExam específico
    user_exam = get_object_or_404(UserExam, id=user_exam_id)

    # Obtener el usuario asociado al UserExam
    usuario = user_exam.usuario

    # Obtener información relevante del examen
    tipo_examen = user_exam.examen.tipo_examen.nombre
    fecha = user_exam.examen.fecha.strftime("%d/%m/%Y")  # Acceder a la fecha del examen
    nota = user_exam.nota  # Usa la nota calculada en el modelo
    aprobado = user_exam.estado  # Usa el estado ya existente

    # Crear el PDF (código para generar el certificado)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="certificado_{usuario.username}.pdf"'

    # Configuración del documento PDF
    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    # Título del certificado
    title = Paragraph("Certificado de Examen", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 0.5 * inch))

    # Información del usuario
    user_info = [
        f"Nombre: {usuario.first_name} {usuario.last_name}",
        f"Cédula: {usuario.cedula}",
        f"Tipo de Examen: {tipo_examen}",
        f"Fecha: {fecha}",
        f"Resultado: {aprobado}",
        f"Nota: {nota}"
    ]

    for info in user_info:
        p = Paragraph(info, styles['Normal'])
        story.append(p)
        story.append(Spacer(1, 0.2 * inch))

    # Agregar una línea al final
    story.append(Spacer(1, 0.5 * inch))
    story.append(Paragraph("______________________________", styles['Normal']))
    story.append(Paragraph("Firma del Administrador", styles['Normal']))

    # Construir el PDF
    doc.build(story)

    return response
